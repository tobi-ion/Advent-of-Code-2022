#https://adventofcode.com/2022/day/4
#a list is given with this format in each cell: 14-98,14-14
#how many assignment pairs does one range fully contain the other

import openpyxl

# Load the workbook and select the first worksheet
workbook = openpyxl.load_workbook('AdventOfCode4.xlsx')
worksheet = workbook.worksheets[0]

# Define the range of cells to compare (starting with row 1)
range_start = 1
range_end = worksheet.max_row

# Initialize the counter for matches
match_count = 0

# Loop through each row of the range
for row in range(range_start, range_end + 1):
    # Get the values in the first column of the row
    values = worksheet.cell(row=row, column=1).value
    # Skip empty cells
    if values is None:
        continue
    print(f"Values: {values}")
    # Convert the values to arrays of integers
    values_list = values.split(',')[0].split('-')
    array1 = [i for i in range(int(values_list[0]), int(values_list[1]) + 1)]
    values_list = values.split(',')[1].split('-')
    array2 = [i for i in range(int(values_list[0]), int(values_list[1]) + 1)]
    #array1 = list(range(int(values.split(',')[0].split('-')[0]), int(values.split(',')[0].split('-')[1])+1))
    #array2 = list(range(int(values.split(',')[1].split('-')[0]), int(values.split(',')[1].split('-')[1])+1))

    # Check if either array is completely within the other
    if set(array1).issubset(set(array2)) or set(array2).issubset(set(array1)):
        match_count += 1

# Print the number of matches found
print(f"Number of matches: {match_count}")


#check for any overlap at all
sum_common = 0  # initialize sum_common to 0

for row in worksheet.iter_rows(min_row=1, max_row=None, min_col=1, max_col=1, values_only=True):
    values = row[0]
    if values is not None:
        values_list1 = values.split(',')[0].split('-')
        values_list2 = values.split(',')[1].split('-')

        array1 = [i for i in range(int(values_list1[0]), int(values_list1[1])+1)]
        array2 = [i for i in range(int(values_list2[0]), int(values_list2[1])+1)]

        if set(array1) & set(array2):
            sum_common += 1

print(sum_common)